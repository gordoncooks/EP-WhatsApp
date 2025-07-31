import pandas as pd
import os
from tkinter import messagebox
from openpyxl import load_workbook

def detect_header_row(df):
    """Detects the most likely header row"""
    for i in range(min(10, len(df))):
        row = df.iloc[i]
        if row.isnull().sum() < len(row) // 2 and all(isinstance(val, str) or pd.isna(val) for val in row):
            return i
    return 0

def read_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    df_raw = pd.read_csv(filepath, header=None) if ext == ".csv" else pd.read_excel(filepath, header=None)
    header_row = detect_header_row(df_raw)
    df = pd.read_csv(filepath, header=header_row) if ext == ".csv" else pd.read_excel(filepath, header=header_row)
    print(f"🔍 Header identified at row {header_row}: {list(df.columns)}")
    return df

def standardize_columns(columns):
    return [str(col).strip().lower().replace(" ", "").replace("#", "").replace("-", "") for col in columns]

def find_matching_columns(df1, df2):
    df1_cols = standardize_columns(df1.columns)
    df2_cols = standardize_columns(df2.columns)

    for i, col1 in enumerate(df1_cols):
        if "listing" in col1:
            for j, col2 in enumerate(df2_cols):
                if "listing" in col2:
                    return df1.columns[i], df2.columns[j]
    return None, None

def clean_listing_column(series):
    return (
        series.astype(str)
        .str.upper()
        .str.replace("P24-", "", regex=False)
        .str.strip()
        .str.split(".")
        .str[0]
    )

def extract_hyperlinks(filepath, target_col):
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        # Find header row containing target_col
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
            headers = [cell.value for cell in row]
            if headers and target_col in headers:
                header_row_idx = i
                break

        if header_row_idx is None:
            print(f"⚠️ Column '{target_col}' not found for hyperlink extraction.")
            return None

        col_idx = headers.index(target_col) + 1
        new_col_index = ws.max_column + 1
        ws.cell(row=header_row_idx, column=new_col_index, value="Extracted Hyperlink")

        hyperlinks = []
        for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row), start=header_row_idx + 1):
            cell = row[col_idx - 1]
            url = cell.hyperlink.target if cell.hyperlink else None
            hyperlinks.append(url)
            ws.cell(row=i, column=new_col_index, value=url)

        wb.close()
        print("🔗 Injected Extracted Hyperlink column into Excel.")
        print(f"🖇️ First 5 extracted hyperlinks: {hyperlinks[:5]}")
        return hyperlinks
    except Exception as e:
        print(f"Failed to extract and inject hyperlinks: {e}")
        return None

def clean_contact_number(num):
    """Cleans and validates a South African phone number"""
    if pd.isna(num) or str(num).strip() == "":
        return None

    # Convert to string and remove common separators
    num = str(num).strip()
    num_cleaned = num.replace(" ", "").replace("-", "").replace(".", "").replace("(", "").replace(")", "")

    # Normalize SA numbers
    if num_cleaned.startswith("0"):
        num_cleaned = "+27" + num_cleaned[1:]
    elif num_cleaned.startswith("+27"):
        num_cleaned = num_cleaned  # already normalized
    else:
        return None

    # Must have at least 11 characters
    if len(num_cleaned) < 11:
        return None

    return num_cleaned

def load_and_match_documents(file1_path, file2_path):
    try:
        print(f"\nAttempting to load files:")
        print(f" - File 1: {file1_path}")
        print(f" - File 2: {file2_path}")

        df1 = read_file(file1_path)
        df2 = read_file(file2_path)
        print("Files loaded successfully.")

        # Try column detection
        p24_col, listing_col = find_matching_columns(df1, df2)

        if not p24_col or not listing_col:
            print("Attempting fallback: swap file roles")
            p24_col, listing_col = find_matching_columns(df2, df1)
            if not p24_col or not listing_col:
                raise ValueError("Matching columns not found in either direction.")
            df1, df2 = df2, df1
            file1_path, file2_path = file2_path, file1_path

        print(f"DATA1 Columns: {list(df1.columns)}")
        print(f"DATA2 Columns: {list(df2.columns)}")
        print(f"Matching '{p24_col}' (DATA1) with '{listing_col}' (DATA2)")

        df1[p24_col] = clean_listing_column(df1[p24_col])
        df2[listing_col] = clean_listing_column(df2[listing_col])

        print("First 5 values in DATA1:", df1[p24_col].dropna().unique()[:5])
        print("First 5 values in DATA2:", df2[listing_col].dropna().unique()[:5])

        # Extract hyperlinks if Excel and applicable
        if file1_path.lower().endswith(".xlsx") and "p24" in p24_col.lower():
            hyperlinks = extract_hyperlinks(file1_path, p24_col)
            if hyperlinks and len(hyperlinks) == len(df1):
                df1["Extracted Hyperlink"] = hyperlinks
            else:
                print("No hyperlinks were extracted or lengths didn't match.")

        merged_df = pd.merge(df1, df2, left_on=p24_col, right_on=listing_col, how="inner")
        print(f"Match complete. {len(merged_df)} records matched.")

        # Clean contact numbers
        contact_col = "Owner Contact Numbers"
        rejected_numbers = []

        if contact_col in merged_df.columns:

            def validate_and_track(num):
                cleaned = clean_contact_number(num)
                if cleaned is None and pd.notna(num) and str(num).strip() != "":
                    rejected_numbers.append(str(num).strip())
                return cleaned

            merged_df[contact_col] = merged_df[contact_col].apply(validate_and_track)

            # Drop rows with no valid numbers
            merged_df = merged_df.dropna(subset=[contact_col]).reset_index(drop=True)
            
            print(f"Rejected numbers: {rejected_numbers}")
        else:
            print("⚠️ 'Owner Contact Numbers' column not found in merged data.")

        return merged_df

    except Exception as e:
        print(f"Error occurred: {e}")
        messagebox.showerror("Merge Failed", f"An error occurred while matching the documents:\n\n{str(e)}")
        return None
